import openpyxl
import os
import time
def UserSplit(string):
    users=list()
    y=string.split('can')
    if 'and' in y:
        y=y[0].rstrip()
    try:y=y.replace(' ','')
    except:y=y[0].replace(' ','')
    users=y.split(',')
    return users

def UserLookUp(user, appname, appno):
    testdir = os.getcwd()+'\\Test Cases\\'
    wb = openpyxl.load_workbook(testdir+appname + ' Test Cases - Feature ' + str(appno) + '.xlsx')
    ws = wb.get_sheet_by_name('Test Account')
    users =  [str(cell.value) for column in ws.iter_cols(min_col=2,max_col=2,
            min_row=2, max_row=ws.max_row) for cell in column]
    email = [str(cell.value) for column in ws.iter_cols(min_col=3,max_col=3,
            min_row=2, max_row=ws.max_row) for cell in column]
    password =  [str(cell.value) for column in ws.iter_cols(min_col=4,max_col=4,
            min_row=2, max_row=ws.max_row) for cell in column]
    for u, e, p in zip(users,email,password):
        if u == user:
            return (e,p)
def Login(browser,testemail,testpass,currentfeat):
    email = browser.find_element_by_id('i0116')
    email.send_keys(testemail)
    browser.find_element_by_id('idSIButton9').click()
    passElem = browser.find_element_by_id('i0118')
    passElem.send_keys(testpass)
    time.sleep(3)
    browser.find_element_by_id('idSIButton9').click()
    time.sleep(1)
    browser.find_element_by_id('idSIButton9').click()
    time.sleep(1) 
    browser.find_element_by_link_text(currentfeat).click()
    time.sleep(3)

def Create(userstories, currentfeat):
    wb = openpyxl.load_workbook(userstories)
    ws = wb.get_sheet_by_name('Metadata')
    library = [cell.value for column in ws.iter_cols(min_col=2,max_col=2,
                min_row=2, max_row=ws.max_row) for cell in column]
    label = [cell.value for column in ws.iter_cols(min_col=4,max_col=4,
                min_row=2, max_row=ws.max_row) for cell in column]
    required = [cell.value for column in ws.iter_cols(min_col=6,max_col=6,
                min_row=2, max_row=ws.max_row) for cell in column]
    field = [cell.value for column in ws.iter_cols(min_col=7,max_col=7,
                min_row=2, max_row=ws.max_row) for cell in column]
    options = [cell.value for column in ws.iter_cols(min_col=9,max_col=9,
                min_row=2, max_row=ws.max_row) for cell in column]
    list1 = []
    list2 = []
    list3 = []
    for a,b,c,d,e in zip (library,label,required,field,options):
        if a == currentfeat and c == 'Y':
            list1.append(b)
            list2.append(d)
            if e==None:
                list3.append('-')
                continue
            list3.append(e)
    return [list1,list2,list3]
lst = []

def Choice(create):
    for i, j, k in zip(*create):
    #print(i,j,k.lstrip().rstrip())
        k=k.split("-")
        if j =='Choice':
            for some in k:
                if some != '':
                    k.append(some.rstrip().lstrip())
    return (k)
print(Choice(Create('LN App Migration - Legal Online User Stories_v2.xlsm', 'Jurisprudence')))
    