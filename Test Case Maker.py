import openpyxl
from openpyxl.workbook import Workbook
from collections import Counter
import os
fileno = 0
files = [f for f in os.listdir('.') if os.path.isfile(f) and '.xls' in f]
for line in files:
    fileno += 1
    print(str(fileno) +". " + line)
testdir = './Test Cases/'
if not os.path.exists(testdir):
    os.makedirs(testdir)
number = input('Please input the file number for the user stories file: ')
userstories = files[int(number)-1]

form = '.xlsx'
nofeat = list()
Lib= list()
rb= openpyxl.load_workbook('.\Templates\RTM Template.xlsx')
rs= rb.get_sheet_by_name('Traceability Matrix')
try: wb = openpyxl.load_workbook(userstories)
except: print('Not a Valid File!') 
listofsheets = wb.get_sheet_names()
try:main_sheet = wb.get_sheet_by_name('User Stories')
except: print('Not a Valid File!') 
#Getting the appname
Dashboard = wb.get_sheet_by_name('Dashboard')
appname = Dashboard['E2'].value
link = Dashboard['E4'].value
perm_sheet = wb.get_sheet_by_name('Permission Matrix')
meta_sheet = wb.get_sheet_by_name('Metadata')
testcase = '.\Templates\Test Cases Template.xlsx'

#Gathering Permission Data
Permission= [cell.value for column in perm_sheet.iter_rows(min_col=2,max_col=perm_sheet.max_column,min_row=1, max_row=1) for cell in column]
LPermission=[cell.value for column in perm_sheet.iter_rows(min_col=1,max_col=perm_sheet.max_column,min_row=2, max_row=perm_sheet.max_row) for cell in column]

# Gather the necessary data
Library =   [cell.value for column in main_sheet.iter_cols(min_col=3,max_col=3,min_row=2, max_row=main_sheet.max_row) for cell in column]
Scenario =  [cell.value for column in main_sheet.iter_cols(min_col=5,max_col=5,min_row=2,max_row=main_sheet.max_row) for cell in column]
Outcome =   [cell.value for column in main_sheet.iter_cols(min_col=6,max_col=6,min_row=2,max_row=main_sheet.max_row) for cell in column]
Features =  [cell.value for column in main_sheet.iter_cols(min_col=1,max_col=1,min_row=2,max_row=main_sheet.max_row) for cell in column]
Users =     [cell.value for column in main_sheet.iter_cols(min_col=4,max_col=4,min_row=2,max_row=main_sheet.max_row) for cell in column]
UScenario = [cell.value for column in main_sheet.iter_cols(min_col=5,max_col=5,min_row=2,max_row=main_sheet.max_row) for cell in column]

#Gather the data for Requirements Tracability Matrix
RTMLibrary =   [cell.value for column in main_sheet.iter_cols(min_col=3,max_col=3,min_row=2, max_row=main_sheet.max_row) for cell in column]
RTMScenario =  [cell.value for column in main_sheet.iter_cols(min_col=5,max_col=5,min_row=2,max_row=main_sheet.max_row) for cell in column]
RTMOutcome =   [cell.value for column in main_sheet.iter_cols(min_col=6,max_col=6,min_row=2,max_row=main_sheet.max_row) for cell in column]
RTMFeatures =  [cell.value for column in main_sheet.iter_cols(min_col=1,max_col=1,min_row=2,max_row=main_sheet.max_row) for cell in column]
RTMUsers =     [cell.value for column in main_sheet.iter_cols(min_col=4,max_col=4,min_row=2,max_row=main_sheet.max_row) for cell in column]
RTMUScenario = [cell.value for column in main_sheet.iter_cols(min_col=5,max_col=5,min_row=2,max_row=main_sheet.max_row) for cell in column]
#Gathering Metadata
LMetadata = [cell.value for column in meta_sheet.iter_cols(min_col=2,max_col=2,min_row=2, max_row=meta_sheet.max_row) for cell in column]
FMetadata = [cell.value for column in meta_sheet.iter_cols(min_col=4,max_col=4,min_row=2, max_row=meta_sheet.max_row) for cell in column]
RMetadata = [cell.value for column in meta_sheet.iter_cols(min_col=6,max_col=6,min_row=2, max_row=meta_sheet.max_row) for cell in column]

totalfeatures = Counter(Library)
for i in totalfeatures.items():
    nofeat.append(i[1])
    Lib.append(i[0])

#print (len(Library), len(Scenario), len(Outcome), len(Features))
#print(len(nofeat))

#=============================================================================
#                     NEXT WORKBOOK
#=============================================================================

testid = 1
feature = 1
count=0
while True:
    metastring = ''
    tb = openpyxl.load_workbook(testcase)
    listofsheets = tb.get_sheet_names()
    mainsheet = tb.get_sheet_by_name(listofsheets[0])
    permsheet = tb.get_sheet_by_name('Permission Matrix')
    testsheet = tb.get_sheet_by_name('Test Account')
    mainsheet['A1'].value = appname + ' Test Cases'
    mainsheet['B1'].value = Library[0]
    mainsheet['C1'].value = link
    featiter= mainsheet.iter_cols(min_col=2, max_col=2, min_row=3,max_row=100) #2+nofeat[feature-1])
    sceniter= mainsheet.iter_cols(min_col=3, max_col=3, min_row=3,max_row=100) #2+nofeat[feature-1])
    conditer= mainsheet.iter_cols(min_col=4, max_col=4, min_row=3,max_row=100) #2+nofeat[feature-1])
    outciter= mainsheet.iter_cols(min_col=7, max_col=7, min_row=3,max_row=100) #2+nofeat[feature-1])
    prociter= mainsheet.iter_cols(min_col=5, max_col=5, min_row=3,max_row=100)
    #Getting the Required Metadata
    for a,b,c in zip(LMetadata,FMetadata,RMetadata):
        if c =='Y' and a==Lib[0]:
            metastring += '-' + b + '\n'
    
    #ITERATION WORKSHEET
    for featiters,sceniters,conditers,outciters,prociters in zip (featiter,sceniter,conditer,outciter,prociter):
        for b,c,d,e,f in zip(featiters,sceniters,conditers,outciters,prociters):
            #Define Library Name
            st = '1. Go to ' + Library[0]
            
            #If the Condition value exists, it will skip
            try: 
                if len(d.value) >0:
                    continue
            #If blank, it will write a new user story
            except:            
                b.value = Features[0]
                c.value = Scenario[0]
                e.value = Outcome[0]
                d.value = Users[0]+' can '+ UScenario[0].lower()
                
            if 'Delete' in str(c.value):
                f.value = '1. Go to ' + Library[0] + '\n2. Select the item to delete.\n3. Click Delete Item.'
            if 'Approve' in str(c.value):
                f.value = '1. Select a document request for approval from the email notification\n2. Review and edit the content\n3. Approve the document'
            if 'Reject' in str(c.value):
                f.value = '1. Select a document request for approval from the email notification\n2. Review and edit the content\n3. Reject the document'
            if 'Create' in str(c.value):
                f.value = st + '\n2. Create new item\n3. Input all necessary details\n4. Submit/Post the new item.'
                prociters[sceniters.index(c)+1].value = st + '\n2. Create new item.\n3. Submit/post with blank values.'
                conditers[sceniters.index(c)+1].value = "Validate required fields"
                outciters[sceniters.index(c)+1].value = "An error message will be displayed on the required fields:\n" + metastring
            if 'Edit' in str(c.value):\
                f.value = st + '\n2. Select and open item to update\n3. Edit the necessary fields.\n4. Submit/post the updated item.'
            
            if 'Upload' in str(c.value): 
                f.value = st + '\n2. Upload new document\n3. Populates all the necessary fields.\n4. Submit/post the new document.'
                prociters[sceniters.index(c)+1].value = st + '\n2. Upload new document\n3. Submit/post with blank values.'
                conditers[sceniters.index(c)+1].value = "Validate required fields"
                outciters[sceniters.index(c)+1].value = "An error message will be displayed on the required fields:\n" + metastring
            # if 'Approve' in str(c.value):continue
            if 'View' in str(c.value):
                f.value = '1. Go to ' + Library[0] + '\n2. Select and open the item to view.'
            if 'Filter' in str(c.value):
                    filterlst=e.value.split('-')
                    for i in range(1, len(filterlst)):
                        prociters[sceniters.index(c) + i - 1].value = st + '\n2. Filter the list by ' + filterlst[i]
                        conditers[sceniters.index(c)+i -1].value = 'Filter the list by ' + filterlst[i]
                        outciters[sceniters.index(c)+i-1].value = "The list of items will be filtered based on the selected criteria."
            del Features[0]
            del Scenario[0]
            del Users[0]
            del UScenario[0]
            del Outcome[0]
            del Library[0]
            
            count+=1
            if nofeat[feature-1] == count: 
                count = 0
                break
            
    #Screenshot and Actual results
    testgen = [cell.value for column in mainsheet.iter_cols(min_col=4,max_col=4,min_row=3,max_row=main_sheet.max_row) for cell in column if cell.value != None]
    for column in mainsheet.iter_cols(min_col=8, max_col=9, min_row=3,max_row=2+len(testgen)): #CHANGE
        for cell in column:
            cell.value = 'Please refer to '+appname+ ' Test Results Appendix - Feature ' + str(feature)+form
    #Test ID Generation
    for column in mainsheet.iter_cols(min_col=1, max_col=1, min_row=3,max_row=2+len(testgen)):
        for cell in column:
            cell.value = "F"+str(feature)+"-"+appname+"-TC-"+str("{0:0=3d}".format(testid))
            testid += 1
    
#==============================================================================  
    #GENERATION OF REQUIREMENTS TRACEABILITY MATRIX
#==============================================================================
    #MATCHING THE TEST CASE IDs to USER STORIES
    testcol=[cell.value for col in mainsheet.iter_cols(min_col=1, max_col=1, min_row=3,max_row=mainsheet.max_row) for cell in col if cell.value != None]
    usercol=[cell.value for col in mainsheet.iter_cols(min_col=2, max_col=2, min_row=3,max_row=2 + len(testcol)) for cell in col]

    teststring = ''
    testtrig = 0
    testlist = list()
    for a1,a2 in zip(testcol,usercol):
        
        if a2 != None and testtrig ==1: 
            testlist.append(teststring)
            teststring = a1 + '\n'
            testtrig =2
            #print(teststring +' '+ str(testtrig) + str(testlist) + 'Went to First')
            continue
        if a2 == None:
            testtrig = 1
            teststring += a1 + '\n'
            #print(teststring +' '+ str(testtrig) + str(testlist)+'Went to Second')
            continue
        if testtrig ==2 and a2 !=None:
            testlist.append(teststring)
            teststring = ''
            testtrig = 2
            teststring += a1 + '\n'
            #print(teststring +' '+ str(testtrig) + str(testlist) +'Went to Third')
            continue
        teststring += a1 + '\n'
        testtrig = 2
        
        #print(teststring +' '+ str(testtrig) + str(testlist) +'Went to Fourth')
    if testtrig == 2:    
        testlist.append(teststring)
    #print(testlist)
    testtrig=0
    #Gathering the Cells for RTM
    featrtm= rs.iter_cols(min_col=1, max_col=1, min_row=rs.max_row,max_row=rs.max_row + nofeat[feature-1]) 
    condrtm= rs.iter_cols(min_col=2, max_col=2, min_row=rs.max_row,max_row=rs.max_row + nofeat[feature-1])
    procrtm= rs.iter_cols(min_col=3, max_col=3, min_row=rs.max_row,max_row=rs.max_row + nofeat[feature-1])
    outcrtm= rs.iter_cols(min_col=4, max_col=4, min_row=rs.max_row,max_row=rs.max_row + nofeat[feature-1]) 
    testrtm= rs.iter_cols(min_col=5, max_col=5, min_row=rs.max_row,max_row=rs.max_row + nofeat[feature-1])
    scenrtm= rs.iter_cols(min_col=6, max_col=6, min_row=rs.max_row,max_row=rs.max_row + nofeat[feature-1])
    #Writing in RTM
    for a1,a2,a3,a4,a5,a6 in zip(featrtm,condrtm,procrtm,outcrtm,testrtm,scenrtm):
        for a,b,c,d,e,f in zip(a1,a2,a3,a4,a5,a6):
                
                a.value = RTMFeatures[0]
                b.value = RTMUsers[0]+' should be able to '+ RTMUScenario[0].lower()
                c.value = 'DFD' + str('{0:0=2d}').format(feature)
                d.value = RTMOutcome[0]
                try: e.value = testlist[0]
                except: print('none printed')
                f.value = RTMScenario[0]
                del RTMFeatures[0]
                del RTMScenario[0]
                del RTMUsers[0]
                del RTMUScenario[0]
                del RTMOutcome[0]
                del RTMLibrary[0]
                try: del testlist[0]
                except: print('none deleted')
                count+=1
                if nofeat[feature-1]==count:
                    count = 0
                    break
                
#==============================================================================
    #TEST CASES
#==============================================================================
    #MERGING CELLS
    mergerow = len(testgen)
    del testgen
    testid = 1
       
    #MERGING CELLS
    trigger=0
    Merge =   [cell.value for column in mainsheet.iter_cols(min_col=3,max_col=3,min_row=3, max_row=2+mergerow) for cell in column]

    
    for i, val in enumerate(Merge):
        if val == None and trigger == 0:
            midx = i-1
            trigger = 1
        if val != None and trigger == 1:
            endidx = i
            trigger = 0
            mainsheet.merge_cells(start_column=3, end_column =3 , start_row =3+ midx, end_row = 2+ endidx)
            mainsheet.merge_cells(start_column=2, end_column =2 , start_row =3+ midx, end_row = 2+ endidx)
    #tb.merge_cells(start_column=3, end_column =3 , start_row =2+ midx, end_row = 2+ endidx)
    tb.save(testdir+appname+' Test Cases - Feature '+ str(feature) + form)
    feature += 1

    if feature > len(nofeat):
        break

#==============================================================================
#             GENERATION OF REQUIREMENTS TRACEABILITY MATRIX
#==============================================================================
rb.save(testdir+'FR-ISDD-001_Requirement Traceability Matrix_'+appname+'.xlsx')