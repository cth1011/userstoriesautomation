#==============================================================================
#      IMPORTING PACKAGES
#==============================================================================
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
import time
import openpyxl
import os
import docx

# Functions
def UserSplit(string):
    users=list()
    usersm=list()
    y=string.split('can')
    if 'and' in y:
        y=y[0].replace('and','')
    try:y=y.rstrip().lstrip()
    except:y=y[0].rstrip().lstrip()
    try:users=y.split(',')
    except:users.append(y)
    for i in users:
        usersm.append(str(i).lstrip().rstrip())
    return usersm

def UserLookUp(user, appname, appno):
    testdir = os.getcwd()+'\\Test Cases\\'
    wb = openpyxl.load_workbook(testdir+appname + ' Test Cases - Feature ' + str(appno) + '.xlsx')
    ws = wb.get_sheet_by_name('Test Account')
    users =  [str(cell.value) for column in ws.iter_cols(min_col=2,max_col=2,
            min_row=2, max_row=ws.max_row) for cell in column]
    email = [str(cell.value) for column in ws.iter_cols(min_col=3,max_col=3,
            min_row=2, max_row=ws.max_row) for cell in column]
    password =  [str(cell.value) for column in ws.iter_cols(min_col=4,max_col=4,
            min_row=2, max_row=ws.max_row) for cell in column]
    for u, e, p in zip(users,email,password):
        if u == user:
            return (e,p)
def Login(browser,testemail,testpass):
    email = browser.find_element_by_id('i0116')
    email.send_keys(testemail)
    browser.find_element_by_id('idSIButton9').click()
    passElem = browser.find_element_by_id('i0118')
    passElem.send_keys(testpass)
    time.sleep(3)
    browser.find_element_by_id('idSIButton9').click()
    time.sleep(1)
    browser.find_element_by_id('idSIButton9').click()
    time.sleep(1)
def Create(userstories, currentfeat):
    wb = openpyxl.load_workbook(userstories)
    ws = wb.get_sheet_by_name('Metadata')
    library = [cell.value for column in ws.iter_cols(min_col=2,max_col=2,
                min_row=2, max_row=ws.max_row) for cell in column]
    label = [cell.value for column in ws.iter_cols(min_col=4,max_col=4,
                min_row=2, max_row=ws.max_row) for cell in column]
    required = [cell.value for column in ws.iter_cols(min_col=6,max_col=6,
                min_row=2, max_row=ws.max_row) for cell in column]
    field = [cell.value for column in ws.iter_cols(min_col=7,max_col=7,
                min_row=2, max_row=ws.max_row) for cell in column]
    options = [cell.value for column in ws.iter_cols(min_col=9,max_col=9,
                min_row=2, max_row=ws.max_row) for cell in column]
    list1 = []
    list2 = []
    list3 = []
    for a,b,c,d,e in zip (library,label,required,field,options):
        if a == currentfeat and c == 'Y':
            list1.append(b)
            list2.append(d)
            if e==None:
                list3.append('-')
                continue
            list3.append(e)
    return [list1,list2,list3]
#==============================================================================
#       Opening the Workbook
#==============================================================================
#files = [f for f in os.listdir('.') if os.path.isfile(f)]
#fileno = 0
appno = 0
#for line in files:
#    fileno += 1
#    print(str(fileno) +". " + line)
#number = input('Please input the file number for the user stories file: ')
userstories = 'LN App Migration - Legal Online User Stories_v2.xlsm'# files[int(number)-1]
testdir = './Test Cases/'
screendir = './Screenshots/'
if not os.path.exists(screendir):
    os.makedirs(screendir)
wb = openpyxl.load_workbook('LN App Migration - Legal Online User Stories_v2.xlsm')#userstories)
ws = wb.get_sheet_by_name('Dashboard')
link = ws['E4'].value
appname = ws['E2'].value

#Opening Browser and Logging In

#==============================================================================
#       Getting the worksheet data
#==============================================================================
while True:
    appno = 2
    appstring = str("{0:0=2d}".format(appno))
    try: wb = openpyxl.load_workbook(testdir+appname + ' Test Cases - Feature '
                                     + str(appno) + '.xlsx')
    except: break
    listofsheets = wb.get_sheet_names()
    mainsheet = wb.get_sheet_by_name(listofsheets[0])
    #Data Gathering from the Test Cases
    currentfeat = mainsheet['B1'].value
    testid =    [cell.value for column in mainsheet.iter_cols(min_col=1,max_col=1,
                min_row=3, max_row=mainsheet.max_row) for cell in column if cell.value !=None]
    condition = [cell.value for column in mainsheet.iter_cols(min_col=4,max_col=4,
                min_row=3, max_row=mainsheet.max_row) for cell in column if cell.value !=None]
    metadata= Create(userstories,currentfeat)
    #Getting Screenshot for Homepage
    if currentfeat == 'Nonthing':#'Landing Page':
        browser= webdriver.Chrome('.\chromedriver.exe')
        browser.get(link)
        browser.get_screenshot_as_file(screendir+testid[0]+'.png')
        print(currentfeat + ' screenshot taken and saved as '+testid[0]+'.png')
        browser.find_element_by_tag_name('html').send_keys(Keys.PAGE_DOWN)
        time.sleep(2)
        browser.quit()
        continue
    #Getting Screenshot for other modules
    for a, b in zip(testid, condition):
        # Defined Function in determing the list of Users
        
        #This is the app no to be added for screenshot formatting
        c = appstring + a
        if 'view' in b:
            users = UserSplit(b)
            for j in users:
                testacct=UserLookUp(j,appname, appno)
                browser= webdriver.Chrome('.\chromedriver.exe')
                browser.get(link)
                Login(browser,testacct[0],testacct[1])
                browser.find_element_by_link_text(currentfeat).click()
                time.sleep(3)
                browser.get_screenshot_as_file(screendir+c+'-01.png')
                print('Printing a screenshot from: ' + currentfeat + '\nFor Test ID: ' + a )
                browser.find_element_by_partial_link_text('Test').click()
                time.sleep(2) 
                browser.get_screenshot_as_file(screendir+c+'-02.png')
                print('Printing a screenshot from: ' + currentfeat + '\nFor Test ID: ' + a )
                browser.close()
        
        if 'Validate' in b:
            users = UserSplit(condition[condition.index(b)-1])
            for j in users:
                testacct=UserLookUp(j,appname, appno)
                browser= webdriver.Chrome('.\chromedriver.exe')
                browser.get(link)  
                Login(browser,testacct[0],testacct[1])
                browser.find_element_by_link_text(currentfeat).click()
                time.sleep(2)
                browser.find_element_by_link_text('new item').click()
                time.sleep(2)
                browser.find_element_by_link_text('Save').click()
                time.sleep(1)
                browser.get_screenshot_as_file(screendir+j+'_'+c+'.png')
                print('Printing a screenshot from: ' + currentfeat + '\nFor Test ID: ' + a )
                browser.close()
        if 'upload' in b:
            users = UserSplit(b)
            for j in users:
                testacct=UserLookUp(j,appname, appno)
                browser= webdriver.Chrome('.\chromedriver.exe')
                browser.get(link)  
                Login(browser,testacct[0],testacct[1])
                browser.find_element_by_link_text(currentfeat).click()
                time.sleep(2)
                browser.find_element_by_link_text('new item').click()
                time.sleep(2)
                for label,ftype,choice in zip(*metadata):
                    print(choice.split('-')[1])
                    try: Elem = browser.find_element_by_xpath("//input[@title='"+label+" Required Field']")
                    except:
                        if ftype == 'Choice':
                            Elem = browser.find_element_by_xpath("//input[@type='checkbox']").click()
                            continue
                        else:
                            print(label + ' was not filled out due to error')

                    if ftype =='Date':
                        Elem.send_keys('12/12/12')
                        continue
                    if ftype =='Choice':
                        Elem.send_keys(choice.split('-')[1].lstrip().rstrip())
                        print(choice.split('-')[1].lstrip().rstrip())
                        continue
                    Elem.send_keys('test')
                    #if ftype =='People or Groups':
                        
            print('Printing a screenshot from: ' + currentfeat + '\nFor Test ID: ' + a )
    break


