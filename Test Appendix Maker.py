import docx
from docx.enum.section import WD_ORIENT
from docx.shared import Inches

import openpyxl
import os
import time
fileno = 0
files = [f for f in os.listdir('.') if os.path.isfile(f) and '.xls' in f]
for line in files:
    fileno += 1
    print(str(fileno) +". " + line)
testdir = './Test Cases/'

number = input('Please input the file number for the user stories file: ')
userstories = files[int(number)-1]
appno = 0
try: wb = openpyxl.load_workbook(userstories)
except: print('Not a Valid File!') 
listofsheets = wb.get_sheet_names()
Dashboard = wb.get_sheet_by_name('Dashboard')
appname = Dashboard['E2'].value
while True:
    appno += 1
    try: wb = openpyxl.load_workbook(testdir+appname + ' Test Cases - Feature ' + str(appno) + '.xlsx')
    except: break
    listofsheets = wb.get_sheet_names()
    mainsheet = wb.get_sheet_by_name(listofsheets[0])
    TestCase = [cell.value for column in mainsheet.iter_cols(min_col=1,max_col=1,min_row=3,max_row=mainsheet.max_row) for cell in column if cell.value != None]
    Outcome = [cell.value for column in mainsheet.iter_cols(min_col=7,max_col=7,min_row=3,max_row=mainsheet.max_row) for cell in column if cell.value != None]
#=============================================================================
#CREATING THE DOCUMENT
#=============================================================================
    doc = docx.Document()
#Changing the Page Orientation through Sections
    sections=doc.sections
    sections[0].orientation = WD_ORIENT.LANDSCAPE
    sections[0].page_width = Inches(11)
    sections[0].page_height = Inches(8.5)
    sections[0].left_margin = Inches(0.16)
    sections[0].right_margin = Inches(0.16)
    sections[0].top_margin = Inches(0.16)
    sections[0].bottom_margin = Inches(0.16)
    
#Adding Table Size
    table = doc.add_table(rows=1 + len(TestCase), cols=4)
    
#Adding the Header Titles
    row = table.rows[0]
    row.cells[0].text = 'TEST CASE'
    row.cells[1].text = 'EXPECTED RESULTS'
    row.cells[2].text = 'ACTUAL RESULTS'
    row.cells[3].text = 'REMARKS'

#Transferring Data
    for i in range(1, len(TestCase)+1):
        row = table.rows[i]
        row.cells[0].text = TestCase[0]
        row.cells[1].text = Outcome[0]
        del TestCase[0]
        del Outcome[0]
#Styling the Table
    table.style = 'Table Grid'
#=============================================================================



#Save the new Document
    doc.save(testdir+'{:s} Test Results Appendix - Feature {:d}.docx'.format(appname, appno))
    print('Saving: ' + testdir+'{:s} Test Results Appendix - Feature {:d}.docx'.format(appname, appno))
time.sleep(5)